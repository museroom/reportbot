import warnings 
import os
from itertools import chain
from django.db.models import Q
import json

#FIXME still need tempfile,zipfile,smart_str?
import tempfile, zipfile
from django.utils.encoding import smart_str
#from django.core.servers.basehttp import FileWrapper

# for handling zip file download
from wsgiref.util import FileWrapper
from StringIO import StringIO
from io import BytesIO
import mimetypes
from django.http import StreamingHttpResponse

# OpenPyXL for generating XLSX
import re
from openpyxl.drawing.image import Image as xlsx_Image
from openpyxl.styles.alignment import Alignment as xlsx_alignment
from openpyxl import load_workbook as xlsx_load_workbook
from openpyxl import Workbook as xlsx_workbook
from openpyxl.styles.borders import Border as xlsx_Border
from openpyxl.styles.borders import Side as xlsx_Side
from wand.image import Image as wand_Image
import xlsxwriter
import hashlib
try:
    from urllib.request import urlopen
except ImportError:
    from urllib2 import urlopen

try:
	import Image
except ImportError:
	from PIL import Image

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin

from django.views.generic.dates import ArchiveIndexView, DateDetailView,  \
	DayArchiveView, MonthArchiveView, \
	YearArchiveView
from django.views.generic import TemplateView
from django.views.generic.detail import DetailView
from django.views.generic.list import ListView
from django.views.generic.base import RedirectView
from django.http import HttpResponseRedirect, HttpResponse
from django.core.urlresolvers import reverse
from django.utils.safestring import mark_safe
from django.utils.text import slugify
from django.utils import timezone
from django.core.files.base import ContentFile
from django.conf import settings

from django.views.generic import View
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt

from .models import Photo, Gallery, DailyReportItem, DepartmentItem, \
	Department, DailyReport, PhotoGroup, Profile, Company, \
	PhotoGroupImage, PhotoGroupImageClass, InventoryType, InventoryItem, \
	InstanceMessage

from django.contrib.auth.forms import UserCreationForm
from django.views.generic.edit import CreateView
from django.views.generic.edit import UpdateView

from .forms import PhotoUploadForm, PhotoGroupPMForm, PhotoGroupCMForm, \
	InstanceMessageForm, InventoryTypeForm, InventoryItemForm

import time, datetime
from HTMLParser import HTMLParser

from utils.logger import logger
from utils.failover import get_department_item_failover

# Json Query.
def JsonPhotoQuery( request, report_item, date_and_time ):
	today = time.strftime( "%y%m%d", time.localtime() ) 
	today = "170706"
	logger( "[JsonPhotoQuery] report_item = {}/ date_and_time = {}".format(
				report_item, date_and_time ) )
	qset = Photo.objects.filter( 
		daily_report_item__pk = int(date_and_time))
	rsp = []
	for thumbSet in qset:
		rsp.append( {"image":thumbSet.get_report_url()} ) 
	#return JsonResponse(list(qset.values('image')), safe=False)
	return JsonResponse(dict(photos=rsp)) #, safe=False)

def JsonTableMapQuery( request, date_and_time ):
	logger( "JsonTableMapQuery {}".format( 
		date_and_time ), 'DEBUG' )
	qset = DailyReportItem.objects.filter( 
							daily_report__title = date_and_time
							).order_by('reportOrder')
	rspGroup = []
	rsp=[]
	for item in qset:
		#rsp = []
		rsp.append( {"order":str(item.reportOrder),
								 "name":item.department_item.name,
								 "pk":str(item.pk),
								 "template":item.tableTemplate,
								 "row_photo":item.photoCol,
								 })
		
		#rsp.append( { "name":item.daily_report_item.name, 
		#rspGroup.append( rsp )
	logger( dict(log=rsp), is_json=True )
	return JsonResponse(dict(tableMap=rsp)) #, safe= False )
		
def JsonReportItemQuery( request, report_pk ):
	logger( "JsonReportItemQuery {}:{}".format( 
		report_pk, 
		DailyReportItem.objects.get( pk=report_pk )
		),
		'DEBUG' 
	)

	HTMLrsp = []
	
	class PrettyHTML( HTMLParser ):
		def handle_starttag( self, tag, attrs ):
			#logger( "start tag:{}".format(tag))
			#logger( "start attrs{}".format(attrs))
			HTMLrsp.append( {"startTag":tag} )
			HTMLrsp.append( {"startAttrs":mark_safe(attrs)} )
		def handle_endtag( self, tag ):
			#logger( "end tag:{}".format(tag))
			HTMLrsp.append( {"endTag":tag} )
		def handle_data( self, data ):
			#logger( "data:{}".format(data))
			HTMLrsp.append( {"data":data} )
			

#	qset = DailyReportItem.objects.filter(
#							daily_report__title = date_and_time
#							).filter( 
#								daily_report_item__name = report_item )
	qset = DailyReportItem.objects.get( pk = report_pk )

	#today = time.strftime( "%y%m%d", time.localtime() ) 
	#today = "170703"
	#qset = DailyReportItem.objects.filter( 
	#										daily_report_item__name = report_item )\
	#										.filter( title__contains = today )
	# parse CKHtml into json

	HTMLstatus = []
	HTMLplan = []
	HTMLstatus_toc = []
	HTMLplan_toc = []
	parser = PrettyHTML()
	
	rsp = []
	rspGroup = []

	HTMLrsp = []
	#logger( "statusCKtext:{}".format(qset.statusCK) )
	qset.statusCK = qset.statusCK.replace("&nbsp;"," ")
	parser.feed( qset.statusCK )
	HTMLstatus = HTMLrsp
	HTMLrsp = []
	qset.planCK = qset.planCK.replace("&nbsp;"," ")
	parser.feed( qset.planCK )
	HTMLplan = HTMLrsp
	HTMLrsp = []
	qset.status_TOC_CK = qset.status_TOC_CK.replace("&nbsp;"," ")
	parser.feed( qset.status_TOC_CK )
	HTMLstatus_toc = HTMLrsp
	HTMLrsp = []
	qset.plan_TOC_CK = qset.plan_TOC_CK.replace("&nbsp;"," ")
	parser.feed( qset.plan_TOC_CK )
	HTMLplan_toc = HTMLrsp

	# if TOC content empty, copy from table HTML text for lazy ass
	if len(HTMLstatus_toc) == 0:
		HTMLstatus_toc = HTMLstatus
	if len(HTMLplan_toc) == 0:
		HTMLplan_toc = HTMLplan

	#rspGroup.append ({
	rspGroup={
		"time_start":str(qset.time_start)[0:5],
		"time_stop" :str(qset.time_stop)[0:5],
		"name":qset.__str__(),
		"statusCK" : HTMLstatus, #,set[0].statusCK,
		"planCK" : HTMLplan, #qset.planCK
		"status_toc_CK": HTMLstatus_toc,
		"plan_toc_CK": HTMLplan_toc,
		"photoCol":qset.photoCol,
		"reportRowID":qset.reportRowID,
		"reportRowID":qset.reportRowID,
		"photoCol":qset.photoCol,
		"reportOrder":qset.reportOrder,
		"rowTableId":qset.rowTableId,
		"colTableId":qset.colTableId,
		"rowItemName":qset.rowItemName,
		"colItemName":qset.colItemName,
		"rowStatus":qset.rowStatus,
		"colStatus":qset.colStatus,
		"rowPlan":qset.rowPlan,
		"colPlan":qset.colPlan,
		"rowTime":qset.rowTime,
		"colTime":qset.colTime,
		"rowDirection":qset.rowDirection,
		"colDirection":qset.colDirection,
		"pk":qset.pk,
		"department":qset.department_item.department.name,
		"daily_report_item":qset.department_item.name,
		"direction":qset.department_item.location,
		"name_long":qset.department_item.name_long, 
		#});
		}
	logRspGroup = rspGroup
#	for field in ('statusCK', 'planCK', 'status_toc_CK',  'plan_toc_CK' ):
		#logRspGroup[field] = mark_safe( rspGroup[field] )
#		logRspGroup[0][field] = mark_safe( rspGroup[0][field] )
#	logRspGroup['statusCK'] = mark_safe( rspGroup['statusCK'] )
#	logRspGroup['planCK'] = mark_safe( rspGroup['planCK'] )
#	logRspGroup['status_toc_CK'] = mark_safe( rspGroup['status_toc_CK'] )
#	logRspGroup['plan_toc_CK'] = mark_safe( rspGroup['plan_toc_CK'] )
	#return JsonResponse(list(qset.values('image')), safe=False)
	#return JsonResponse(HTMLrsp, safe=False)
	logger( logRspGroup, is_json=True )
	return JsonResponse(dict(reportItems=rspGroup) ) #, safe=False)


# ReportItem / Daily Report Views
class DailyReportDateView(object):
	queryset = Photo.objects.all()
	date_field = 'date_added'
	allow_empty = True


# PhotoGroup / Monthly Report Views
class PhotoGroupDetailView(object):
	queryset = PhotoGroup.objects.all()
	date_field = 'date_added'
	allow_empty = True

# Create PhotoGroup with single image
def Create_PhotoGroup( request, photo_pk ):
	print( "create photogroup: request={} photo_pk={}".format(
	       request, photo_pk ) )
	q_photo = Photo.objects.get( pk = photo_pk )
	photo_date_time = q_photo.date_added.astimezone( timezone.get_default_timezone()) 
	# look for duplicated photogroup before create new case
	qset_pg = PhotoGroup.objects.filter( photo_records__photo = q_photo )
	if qset_pg:
		print( "duplicated PhotoGroup, return exisiting instead" )
		pg_pk = qset_pg[0].pk 
		pg_url = reverse( "photologue:monthly-report-detail", kwargs={'pk':pg_pk} )
		return redirect( pg_url )
	else: 
		# create new case (photogroup)
		if q_photo.department_item:
			q_department_item = q_photo.department_item
		else:
			q_department_item = get_department_item_failover() 
		print( "q_department_item:{}".format( q_department_item))
		groupname = u"{2}_{0}_{1}_auto".format(photo_date_time.strftime( "%y%m%d-%H%M%S" ),
											q_department_item.name,
											q_department_item.department.company.name )
		new_photo_group = PhotoGroup( name=groupname, date_added=photo_date_time )
		new_photo_group.department_item = q_photo.department_item
		new_photo_group.date_of_service = photo_date_time
		new_photo_group.serviced_date = photo_date_time
		new_photo_group.inspection_date = photo_date_time
		new_photo_group.save()
		photogroup_image_class_center = PhotoGroupImageClass.objects.get( name="Center" )
		photogroup_image = PhotoGroupImage(  
							   photo = q_photo,
							   photo_class = photogroup_image_class_center )
		photogroup_image.save() 
		new_photo_group.photo_records.add( photogroup_image )

		return redirect( reverse('photologue:monthly-report-detail', args=[new_photo_group.pk] ))

def Set_dbField_PhotoGroup( request, record_type, photogroup_id, **kwargs ):
	print( "record_type={} photogroup_id={}".format( record_type, photogroup_id ) )
	q_photogroup = PhotoGroup.objects.get( pk = photogroup_id )
	#FIXME validation? probably an UpdateView?
	q_photogroup.record_type = record_type.upper()
	q_photogroup.save()
	return redirect( reverse( 'photologue:monthly-report-detail', kwargs={'pk':photogroup_id} ) )

# Photo Select Pop Views

class PhotoSelectListView(ListView): 
	model = Photo,
	template_name = "photologue/photoselect_list.html"
	
	def get_context_data( self, **kwargs):
		context = super( PhotoSelectListView, self).get_context_data(**kwargs)
		profile = self.request.user.profile
		date_time = timezone.datetime.strptime(  
						"{}-{}-{}".format(
							self.kwargs['year'], 
							self.kwargs['month'],
							self.kwargs['day'] ),"%Y-%m-%d")
		date_time_prev = date_time - timezone.timedelta(1,0,0)
		date_time_next = date_time + timezone.timedelta(1,0,0)
		target = self.kwargs['target']
		pk = self.kwargs['pk']
		context['search_field_url'] = reverse( 'photologue:photo-select-popup-list',
											kwargs={'year':date_time.year,
													'month':date_time.month,
													'day':date_time.day,
													'target':target,
													'pk':pk} )
		context['date_prev'] = date_time_prev
		context['date_prev_url'] = reverse( 'photologue:photo-select-popup-list',
											kwargs={'year':date_time_prev.year,
													'month':date_time_prev.month,
													'day':date_time_prev.day,
													'target':target,
													'pk':pk} )

		context['date_report'] = date_time
		context['date_next'] = date_time_next
		context['date_next_url'] = reverse( 'photologue:photo-select-popup-list',
											kwargs={'year':date_time_next.year,
													'month':date_time_next.month,
													'day':date_time_next.day,
													'target':target,
													'pk':pk} )
		context['active_photogroup']=self.request.user.profile.active_report
		if target == 'photogroup':
			context['target_photo_group'] = target
		elif target == 'dailyreport':
			context['target_daily_report'] = target
		context['pk'] = pk
		user_profile = self.request.user.profile
		context['department_item_list'] = DepartmentItem.objects.filter(
				department__company = user_profile.company )
		context['daily_report_item_list'] = DailyReportItem.objects.filter(
				department_item__department__company = 
				self.request.user.profile.company ).filter( 
						daily_report = user_profile.active_report )

		return context

	def get_queryset(self):
		q_get = self.request.GET.get('q',None)
		if q_get:
			qset_prev = []
			qset = []
			qset_filter = q_get.split(' ')

			print( 'debug: qset_filter={}'.format(qset_filter))
			counter = 0
			for q_filter in qset_filter: 
				qset_rt= Photo.objects.filter( 
			           title__icontains = q_filter )
				qset_n = Photo.objects.filter (
					   tags__icontains = q_filter )
				qset_din = Photo.objects.filter (
					   department_item__name__icontains = q_filter )
				qset.append( list(chain(qset_rt,qset_n,qset_din)))

			qset_tmp = list(chain(qset))[0]
			qset_tmp = qset
			qset = qset_tmp[0]
			for i in range(0,len(qset_tmp)):
				qset = list(set(qset) & set(qset_tmp[i]))
		else: 
			date_time = timezone.datetime.strptime(  
							"{}-{}-{}".format(
								self.kwargs['year'], 
								self.kwargs['month'],
								self.kwargs['day'] ),"%Y-%m-%d")
			qset = Photo.objects.filter( date_added__date = date_time.date() )

		return qset

# CM / PM monthly report

class MonthlyReportListView( LoginRequiredMixin, ListView ):
	login_url = '/login'
	model = PhotoGroup
	template_name="photologue/monthlyreport_list.html"

	def get_context_data( self, **kwargs ):
		context = super( MonthlyReportListView, self).get_context_data(**kwargs) 
		context['search_field_url'] = reverse( 'photologue:monthly-report-list' )
		return context

	def get_queryset(self):
		q_get = self.request.GET.get('q',None)
		if q_get:
			qset_prev = []
			qset = []
			qset_filter = q_get.split(' ')
			print( 'debug: qset_filter={}'.format(qset_filter))
			counter = 0
			for q_filter in qset_filter: 
				qset_rt= PhotoGroup.objects.filter( 
			           record_type__icontains = q_filter ).order_by( 'date_of_service' )
				qset_n = PhotoGroup.objects.filter (
					   name__icontains = q_filter ).order_by( 'date_of_service' )
				qset_din = PhotoGroup.objects.filter (
					   department_item__name__icontains = q_filter ).order_by( 'date_of_service' )
				qset.append( list(chain(qset_rt,qset_n,qset_din)))

			qset_tmp = list(chain(qset))[0]
			qset_tmp = qset
			qset = qset_tmp[0]
			for i in range(0,len(qset_tmp)):
				qset = list(set(qset) & set(qset_tmp[i]))
		else: 
			qset = PhotoGroup.objects.all().order_by(
				"date_of_service" )
		return qset


class MonthlyReportDetailView(LoginRequiredMixin, DetailView ):
	login_url = '/login' 
	model = PhotoGroup
	template_name="photologue/monthlyreport_detail.html"
	def get_context_data( self, **kwargs):
		obj = kwargs['object']
		print( "debug: MonthlyReportDetailView:Obj:photogroup={}".format(obj) )
		q_profile = Profile.objects.get( pk = self.request.user.profile.pk )
		q_profile.active_photogroup = obj 
		q_profile.save()
		date_time = obj.date_added
		context = super( MonthlyReportDetailView, self).get_context_data(**kwargs)
		context['var1'] = 'value1'
		context['add_photo_url'] = reverse( 'photologue:photo-select-popup-list',
					kwargs={'year':date_time.year, 'month':date_time.month, 'day':date_time.day,
							'target':'photogroup', 'pk':obj.pk} )
		context['admin_record_url'] = reverse( 'admin:photologue_photogroup_change', args=[obj.id] ) 
		#context['edit_record_url'] = reverse( 'admin:photologue_photogroup_change', args=[obj.id] )
		context['set_cm_url'] = reverse( 'photologue:set-photogroup-record-type', 
		                                    kwargs={'photogroup_id':obj.id, 'record_type':'cm' } )
		context['set_pm_url'] = reverse( 'photologue:set-photogroup-record-type', 
		                                    kwargs={'photogroup_id':obj.id, 'record_type':'pm' } )
		context['set_active_photo_group_url'] = reverse( 'photologue:set-active-photogroup', 
		                                                 args=[obj.id] )
		if "PM" in str(obj.record_type).upper():
			context['edit_record_url'] = reverse( 'photologue:photogroup-pm-edit', args=[obj.id] )
		else:
			context['edit_record_url'] = reverse( 'photologue:photogroup-cm-edit', args=[obj.id] )
		context['generate_xlsx_url'] = reverse( 'photologue:generate-xlsx', args=[obj.id] )
		q_profile = Profile.objects.get( pk = self.request.user.profile.pk )
		q_profile.active_photogroup = obj
		q_profile.save()
		context['active_photogroup'] = q_profile.active_photogroup
		context['select_company'] = Company.objects.all()
		context['select_department'] = Department.objects.all()
		context['target_photo_group'] = self
		return context

# Daily Report 

class DailyReportListView(DailyReportDateView, ArchiveIndexView ):
	#date_and_time = timezone.localtime().strftime("%y%m%d-%H%M")
	#paginate_by = 5

	template_name = "photologue/dailyreport_edit.html"

	def get_queryset(self):
		print( 'reportitemlistview queryset={0}'.format( self.kwargs ) )
		if self.kwargs.has_key( 'date_and_time' ):
			print( 'reportitemlistview dateandtime={0}'.format( self.kwargs['date_and_time'] ) )
			date_and_time = self.kwargs['date_and_time']
		else:
			print( 'dateandtime set to today' )
			date_and_time = '2017-07-26-1930'

		print( "date_and_time = {}".format(date_and_time))
			
		#return Department.objects.all()
		d = Photo
		qset = d.objects.filter( daily_report_item__daily_report__title = date_and_time ).order_by( 'daily_report_item__reportOrder' ) 
#		return { 'object_list' : qset } #, 'date_and_time':date_and_time }
		return qset

#class DailyReportListView(ListView):
#	queryset = DailyReport.objects.all()
#	paginate_by = 20

class DailyReportDetailView( DetailView):
	queryset = DailyReport.objects.all()


class DailyReportDetailView(DailyReportDateView, DateDetailView):
	pass


class DailyReportArchiveIndexView(DailyReportDateView, ArchiveIndexView):
	pass


class DailyReportDayArchiveView(LoginRequiredMixin, DailyReportDateView, DayArchiveView):

	login_url = '/login/'
	template_name = "photologue/dailyreport_edit.html"
	date_and_time = timezone.localtime()

	def get_context_data( self, **kwargs):
		print("debug: active report={}".format(self.request.user.profile.active_report))
		context = super(DayArchiveView, self).get_context_data(**kwargs)
		context['daily_report'] = DailyReport.objects.all()
		context['active_report'] = self.request.user.profile.active_report.title
		if self.kwargs.has_key('year'):
			print( "debug: dailyrepot_edit using redirect date" )
			date_and_time = "{}-{}-{}-1930".format(
								self.kwargs['year'],
								self.kwargs['month'],
								self.kwargs['day'] )
			report_dt = timezone.make_aware(timezone.datetime.strptime( "{}-{}-{}".format( 
								self.kwargs['year'], self.kwargs['month'], self.kwargs['day']),
								"%Y-%m-%d" ))
			#context['is_active'] = True
		else: # Just return active report 
			report_dt = self.request.user.profile.active_report.report_date.astimezone(
								timezone.get_default_timezone())
			date_and_time = report_dt.strftime( "%Y-%m-%d-%H%M" ) 
			#context['is_active'] = False
		if DailyReport.objects.get( 
				report_date__date = report_dt.date() ) == self.request.user.profile.active_report:
			context['is_active'] = True
		else:
			context['is_active'] = False
					
		print("date_and_time={}".format(report_dt))

		qset = Photo.objects.filter( 
			daily_report_item__daily_report__report_date__date = report_dt.date()).order_by(
					'daily_report_item__reportOrder' )
		context['photo_list'] = qset
		print( "debug: dailyreport_edit report_dt={}".format( report_dt ) )
		qset = DailyReportItem.objects.filter( daily_report__report_date__date = report_dt.date()).order_by(
					'reportOrder' )
		context['daily_report_item_list'] = qset 
		context['select_report_url'] = reverse( 
		                                   'photologue:dailyreport-edit', kwargs={
		                                   'year':report_dt.year,'month':report_dt.month,'day':report_dt.day} )
		context['report_date_time'] = report_dt.strftime( "%Y-%m-%d" )
		#FIXME include 1530/1930 in time please
		context['select_date_time_current'] = report_dt.strftime( "%y%m%d-1930" )
		return context
	
	def dispatch( self, request, *args, **kwargs ):
		print( 'request.GET:{}'.format( request.GET ) )
		select_report = self.request.GET.get('select_report', None )
		print( 'select_report:{}'.format(select_report) )
		if select_report:
			date_time = timezone.datetime.strptime( select_report, "%y%m%d-%H%M" ) 
			print( 'date_time:{}'.format(date_time))
			return redirect( reverse('photologue:dailyreport-edit', kwargs={
			                 'year':date_time.year,'month':date_time.month,'day':date_time.day} ))
		return super(DailyReportDayArchiveView, self).dispatch(request, *args, **kwargs)

	#def get_queryset( self, **kwargs ):
	#	print( 'dailyreportedit:get_queryset' )

	#	return qset

class DailyReportMonthArchiveView(DailyReportDateView, MonthArchiveView):
	pass 

class DailyReportYearArchiveView(DailyReportDateView, YearArchiveView):
	make_object_list = True

# Update Forms

def Update_PhotoGroup( request, photo_group_pk ):
	print( u"DEBUG: photo_group_pk = {}".format( photo_group_pk ) )
	print( u"DEBUG: active_photogroup = {}".format( request.user.profile.active_photogroup ) )
	q_photogroup = request.user.profile.active_photogroup
	photogroup_image_class_center = PhotoGroupImageClass.objects.get( name = "Center" )
	#FIXME delete not working
	for q_photo_pk in request.POST.getlist('add_photo'):
		print( u"adding q_photo_pk={}".format( q_photo_pk ) )
		q_photo = Photo.objects.get( pk = q_photo_pk )
		photogroup_image = PhotoGroupImage( 
		                       photo = q_photo,
							   photo_class = photogroup_image_class_center )
		photogroup_image.save()
		q_photogroup.photo_records.add( photogroup_image )
	for q_photo_pk in request.POST.getlist('delete_photo'):
		print( u"deleting q_photo_pk={}".format( q_photo_pk))
		#q_photo = Photo.objects.get( pk = q_photo_pk )
		q_photogroup_image = PhotoGroupImage.objects.get( pk = q_photo_pk )
		q_photogroup.photo_records.remove( q_photogroup_image )
	
	if request.POST.has_key('delete_photo'):
		return HttpResponseRedirect( reverse(
			'photologue:monthly-report-detail', args=[ photo_group_pk ] ) )
	else:
		return HttpResponseRedirect( reverse( 'photologue:message-success' ) )
	
def Update_DailyReportItem( request, daily_report_pk=None, daily_report_item_pk=None ):
	def is_number( s ):
		try:
			float( s )
			return true
		except ValueError:
			return false

	print( u"DEBUG: daily_report_pk = {} daily_report_item_pk = {} // request.POST= {}".format (
			 daily_report_pk, daily_report_item_pk, request.POST.getlist('report_photo'))) 
#			 request.POST.getlist('department_pk')))
	redirect_url = reverse( 'photologue:report_item_list_view' )
	#FIXME Daily Report temp fix on the bus, but the problem is DEL and ADD not related to photo.pk
	if daily_report_item_pk:
		if '2017' not in daily_report_item_pk:
		#if is_number( daily_report_item_pk ):
	#if not isinstance(daily_report_item_pk, basestring):
	#if daily_report_item_pk:
#		date_time = timezone.datetime.strptime( 
#				"%Y-%m-%d-%H%M", daily_report_item_pk )
			print( 'inside 2017 FIXME' );
			q_dri = DailyReportItem.objects.get( pk = daily_report_item_pk )
			date_time = q_dri.daily_report.report_date
			redirect_url = reverse( 'photologue:photo-select-popup-list',
					kwargs = {'year':date_time.year,
							  'month':date_time.month,
							  'day':date_time.day,
							  'target':'dailyreport',
							  'pk':q_dri.pk }
					)
			for q_photo_pk in request.POST.getlist('add_photo' ):
				q_photo = Photo.objects.get( pk = int(q_photo_pk) )
				q_photo.daily_report_item.add( q_dri )
				q_photo.department_item = q_dri.department_item
				q_photo.save()

		#FIXME stupid get_related_daily_report_item()
		for q_photo_pk in request.POST.getlist('report_photo'):
			print( u"updating pk:{} related daily_report_item".format(q_photo_pk) )
			q_photo = Photo.objects.get( pk = int(q_photo_pk) )
			q_related_daily_report_item = q_photo.get_related_daily_report_item()
			q_photo.daily_report_item.add( q_related_daily_report_item )
			q_photo.save()

	else:
		for form_pk in request.POST.getlist('report_photo'): 
			dri_pk, photo_pk = form_pk.split(  '-' )
			q_photo = Photo.objects.get( pk = photo_pk )
			q_dri = DailyReportItem.objects.get( pk=dri_pk )
			dr_active = request.user.profile.active_report
			qc_dri = DailyReportItem.objects.filter(
				daily_report = dr_active).get(
				department_item = q_dri.department_item ) 
			q_photo.daily_report_item.add( qc_dri )
			print( "[clone] {}>{}".format( qc_dri, q_photo ) )
			redirect_url = reverse( 'photologue:report_item_list_view' )
		for form_pk in request.POST.getlist('delete_photo'):
			dri_pk, photo_pk = form_pk.split( '-' ) 
			q_dri = DailyReportItem.objects.get( pk=dri_pk)
			q_photo = Photo.objects.get( pk=photo_pk )
			q_photo.daily_report_item.remove( q_dri )
			print( '[remove] {}:{}'.format(
				q_dri, q_photo ) ) 
		
	return HttpResponseRedirect( redirect_url )

class PhotoCatagorize(ListView):
	template_name = "photologue/photo_catagorize.html"

	def get_context_data( self, **kwargs ):
		context = super( PhotoCatagorize, self).get_context_data(**kwargs)
		dt_now = timezone.localtime()
		year = self.kwargs.get( 'year', dt_now.year )
		month = self.kwargs.get( 'month', dt_now.month )
		day = self.kwargs.get( 'day', dt_now.day )
		date_current = timezone.datetime.strptime(
		                   "{}-{}-{}".format(
						       year, month, day),
						   "%Y-%m-%d") 
		date_prev = date_current - timezone.timedelta(1,0,0)
		date_next = date_current + timezone.timedelta(1,0,0)
		url_prev = reverse( "photologue:photo_catagorize_date", kwargs={
		                    'year': date_prev.year,
							'month': date_prev.month,
							'day': date_prev.day } )
		url_next = reverse( "photologue:photo_catagorize_date", kwargs={
		                    'year': date_next.year,
							'month': date_next.month,
							'day': date_next.day } ) 
		context['url_date_prev'] = url_prev
		context['url_date_next'] = url_next
		context['date_prev'] = date_prev
		context['date_next'] = date_next
		url_search = reverse( "photologue:photo_catagorize_date", kwargs={
		                    'year': date_current.year,
							'month': date_current.month,
							'day': date_current.day } ) 
		context['search_field_url'] = url_search 

		return context
		

	def get_queryset(self):
		q_get = self.request.GET.get('q',None)
		if q_get:
			qset_prev = []
			qset = []
			qset_filter = q_get.split(' ')
			print( 'debug: qset_filter={}'.format(qset_filter))
			counter = 0
			for q_filter in qset_filter: 
				qset_rt= Photo.objects.filter( 
			           title__icontains = q_filter )
				qset_n = Photo.objects.filter (
					   tags__icontains = q_filter )
				qset_din = Photo.objects.filter (
					   department_item__name__icontains = q_filter )
				qset.append( list(chain(qset_rt,qset_n,qset_din)))

			qset_tmp = list(chain(qset))[0]
			qset_tmp = qset
			qset = qset_tmp[0]
			for i in range(0,len(qset_tmp)):
				qset = list(set(qset) & set(qset_tmp[i]))
		else:
			dt_now = timezone.localtime()
			year = self.kwargs.get( 'year', dt_now.year )
			month = self.kwargs.get( 'month', dt_now.month )
			day = self.kwargs.get( 'day', dt_now.day )
			date_photo = timezone.datetime.strptime( 
						 "{}-{}-{}".format( year, month, day ), 
						 "%Y-%m-%d" )
			qset = Photo.objects.filter( date_added__date = date_photo )

		return qset
	
def SetPhotoDepartmentItem( request ):
	try:
		department_pk = request.POST.getlist('department_pk')[0]
	except:
		return HttpResponseRedirect( reverse( 'photologue:photo_catagorize' ) )
	q_department_item = DepartmentItem.objects.get( pk = department_pk )
	print( "set_dailyreportitem:department={}".format( department_pk ) )
	try:
		list_set_photo = request.POST.getlist('set_photo')
	except:
		return HttpResponseRedirect( reverse( 'photologue:photo_catagorize' ) )
	for set_photo in list_set_photo:
		print( "set_dailyreportitem:{}".format( set_photo))
		q_photo = Photo.objects.get( pk = set_photo )
		q_photo.department_item = q_department_item
		q_photo.save()
	return HttpResponseRedirect( reverse( 'photologue:photo_catagorize' ) )

#def ReportItemListView( request):
#	print( "reportitemlistview called" )
#	queryset = ReportItem.objects.all()
##	paginate_by = 20
#	context = {'qset_report_item': queryset}
#	return render( request, 'photologue/reportitem_list.html',context )
def photosave( filename, content ):
	print("photosave: photo begin")
	p = Photo( title=filename, 
				slug=slugify(filename) )
	print("photosave: photo image save")

# Instance Message (IM) / Forum / media files

#FIXME use class instead of def
@csrf_exempt
def InstanceMessageCreate( request ):
	msg = u"InstanceMessageCreate: request.POST=[{}]".format( 
		request.POST )
	msg_content = u""
	try:
		msg_content = unicode( request.POST['content'] )
	except:
		print( "[error] msg_content" )

	print( u"{}\ncontent:{}".format(msg, msg_content  ) )
	
	#FIXME validate the form

	# store to database
	post = request.POST
	try:
		im = InstanceMessage( msg_id = post['msg_id'],
							  content = post['content'],
							  sender = post['sender'],
							  receiver = post['receiver'],
							  room = post['room'],
							  media = post['media'], )
		im.save() 
	except:
		print( "[error] im db create" )

	return HttpResponse( 'success {}'.format( response ) )

@csrf_exempt
def PhotoUploadView( request ):

	#return HttpResponseRedirect( reverse( 'admin:login' ))
	if request.method == 'POST':
		form = PhotoUploadForm(request.POST, request.FILES)
		if form.is_valid():
			form.save(request=request)
			
			print( "photo finish" )
		else:
			logger( "photouploadview: form.is_valid() false" )
	return render( request, 'photologue/photo_upload.html', {
			'title': 'Photo Upload View',
			'name2': 'test_form2',
		})
	#return HttpResponse( "Success" )

class MonthlyReportPhotoReorder( DetailView ):
	model = PhotoGroup
	template_name='photologue/test-sortable.html'
	
	def get_context_data( self, **kwargs ):
		context = super( MonthlyReportPhotoReorder, self).get_context_data(**kwargs)
		qset_before = self.object.photo_records.filter( 
		                  photo_class__name = "Before" )
		qset_center = self.object.photo_records.filter( 
		                  photo_class__name = "Center" )
		qset_after = self.object.photo_records.filter( 
		                  photo_class__name = "After" )
		print( "{} {} {}".format( len(qset_before), len(qset_center), len(qset_after) ) )
		context['qset_before'] = qset_before
		context['qset_center'] = qset_center
		context['qset_after']  = qset_after 

		return context

def SortableSubmitTest( request, photo_group_pk ):
	q_photo_group = PhotoGroup.objects.get( pk = photo_group_pk )
#	q_photo_group.photo_records.clear()
	for q_photogroup_img in q_photo_group.photo_records.all():
		print( "debug: delete {} from {}".format( 
		       q_photogroup_img, q_photo_group ) )
		q_photogroup_img.delete()
	i = 0
	for q_photo_pk in request.POST.getlist('photo_order'):
		if '/' in q_photo_pk:
			i = i + 1
			try:
				pg_class_name, pg_style, p_pk = q_photo_pk.split('/')
				print( "{}_{}_{}".format( pg_class_name, pg_style, p_pk ) )
				pg_class = PhotoGroupImageClass.objects.get( name = pg_class_name )
				q_photo = Photo.objects.get( pk = p_pk )
				pg_img = PhotoGroupImage( photo = q_photo, photo_class = pg_class )
				pg_img.save()
				q_photo_group.photo_records.add( pg_img )
			except:
				print( "ERROR: dropped unsupported hidden field:"+ q_photo_pk )

	return HttpResponseRedirect( 
			  #reverse( 'photologue:test-sortable', 
			  reverse( 'photologue:monthly-report-detail', 
			  args=[photo_group_pk]) 
			   )
	
# XLSX output with openpyxl
def GenerateXLSX( request, photo_group_pk ):

	print( "-===-===-==-" )
	print( "GenerateXLSX {}".format( photo_group_pk ) )

	q_pg = PhotoGroup.objects.get( id = photo_group_pk )

	static_root = getattr( settings, 'STATIC_ROOT', '' )
	static_url = getattr( settings, 'STATIC_URL', '' )
	if request.is_secure():
		app_url_prefix = "https://"
	else:
		app_url_prefix = "http://"
	app_url = app_url_prefix + request.get_host()
	tmp_root = '/media/djmedia/mr_forgot/tmp'
	xlsx_root = 'xlsx'

	def create_image_data( url ):

		with wand_Image( filename = url ) as wandImg:
			wandImg.format = 'jpg'
			tmp_name = os.path.join( tmp_root,
			hashlib.sha1(url).hexdigest() + '-wand.' + wandImg.format ) 
			wandImg.save( filename=tmp_name ) 

		image_data = BytesIO(open(tmp_name).read())
		return image_data

	def create_image( url, width=None, height=None ):

		#image_data = BytesIO(open(tmp_name).read())
		image_data = create_image_data( url )
		image = xlsx_Image( image_data )
		if width:
			org_width = image.drawing.width
			org_height = image.drawing.height
			image.drawing.width = width
			image.drawing.height = float(org_height) / float(org_width) * width
		if height:
			org_width = image.drawing.width
			org_height = image.drawing.height
			image.drawing.height = height
			image.drawing.width = float(org_width) / float(org_height) * height
		return image


	if 'PM' in str(q_pg.record_type).upper():
		logo_url = app_url + "/media/photologue/photos/logo-pm.png"
		sig_url  = app_url + "/media/photologue/photos/image2_ZJL3Hw2.jpeg"
		filename_in = 'pm-template.xlsx'
	else: # FIXME assume CM 
		logo_url = app_url + "/media/photologue/photos/image1.png"
		filename_in = 'cm-template.xlsx'
	filename_config = 'sc-config.json'
	pg_serial = PhotoGroup.objects.get( pk=photo_group_pk ).serial_no 
	if pg_serial != "":
		filename_serial = PhotoGroup.objects.get( pk=photo_group_pk ).serial_no
	else:
		filename_serial = 'no_serial-{}-'.format( photo_group_pk )
	filename_photo_out = '{}-photo.xlsx'.format( filename_serial )
	filename_text_out = '{}-report.xlsx'.format( filename_serial )
	filename_writer_out = '{}-photo.xlsx'.format( filename_serial )
	filename_zip_out = '{}.zip'.format( filename_serial )
	fn_in = os.path.join(static_root,xlsx_root,filename_in)
	url_in = app_url + os.path.join( static_url, xlsx_root,filename_in )
	fn_text_out = os.path.join(tmp_root,xlsx_root,filename_text_out)
	fn_photo_out = os.path.join(tmp_root,xlsx_root,filename_photo_out)
	fn_writer_out = os.path.join(tmp_root,xlsx_root,filename_writer_out)
	fn_config = os.path.join(static_root,xlsx_root,filename_config)
	
	print( 'url_in='+url_in)
	xlsx_data = BytesIO(urlopen(url_in).read())

	wb_text = xlsx_load_workbook( xlsx_data )
	wb_photo = xlsx_workbook()
	ws = wb_text.active
	wb_writer = xlsxwriter.Workbook( fn_writer_out )

	# hand coded template photo grid value
	if 'PM' in str(q_pg.record_type).upper():
		# XlsxWriter
		ws_photo = {
			'scale':0.125, 'x_scale':0.125, 'y_scale':0.125, 'gap':1, 
			'page_row': 51, 'page_col': 51, 'cell':{'width': 2, 'height':20},
			'row_begin': 5, 'col':{ 'Before': 'B', 'Center': 'D', 'After': 'F'},
			'logo':{'height':70,'record':'C','photo':'C','x_scale':0.5,'y_scale':0.5},
		}
	else: # FIXME assume CM 
		# XlsxWriter
		ws_photo = {
			'scale':0.125, 'x_scale':0.125, 'y_scale':0.125, 'gap':1, 
			'page_row': 51, 'page_col': 51, 'cell':{'width': 2, 'height':20},
			'row_begin': 9, 'col':{ 'Before': 'B', 'Center': 'D', 'After': 'F'},
			'logo':{'height':150,'record':'A','photo':'A','x_scale':0.5,'y_scale':0.5},
		}
	# command ws_photo keys and values
	ws_photo_common = {
			'paper':9, 'margin':{'left':0.6,'right':0.4,'top':0.6,'bottom':0.6},
			'xoffset':{ 'Before': 10, 'Center': 25, 'After': 30 },
		}
	ws_photo.update( ws_photo_common )

	fn_out_path, filename = os.path.split(fn_text_out)
	print( "fn_out_path={}, filename={}".format( fn_out_path, filename ) )
	if not os.path.exists( fn_out_path ): 
		os.mkdir( fn_out_path )

	# Insert Logo on every page
	#ws.add_image( create_image( 
	#	logo_url, 
	#	ws['A1'].anchor[0], ws['A1'].anchor[1], 
	#	height=170 ) )

	fields_pg = q_pg._meta.get_fields()

	# Replace template tag  with database value
	def style_range(ws, cell_range, border=xlsx_Border(), fill=None, font=None, alignment=None):
		"""
		Apply styles to a range of cells as if they were a single cell.

		:param ws:  Excel worksheet instance
		:param range: An excel range to style (e.g. A1:F20)
		:param border: An openpyxl xlsx_Border :param fill: An openpyxl PatternFill or GradientFill
		:param font: An openpyxl Font object
		"""

		top = xlsx_Border(top=border.top)
		left = xlsx_Border(left=border.left)
		right = xlsx_Border(right=border.right)
		bottom = xlsx_Border(bottom=border.bottom)

		first_cell = ws[cell_range.split(":")[0]]
		if alignment:
			ws.merge_cells(cell_range)
			first_cell.alignment = alignment

		rows = ws[cell_range]
		if font:
			first_cell.font = font

		for cell in rows[0]:
			cell.border = cell.border + top
		for cell in rows[-1]:
			cell.border = cell.border + bottom

		for row in rows:
			l = row[0]
			r = row[-1]
			l.border = l.border + left
			r.border = r.border + right
			if fill:
				for c in row:
					c.fill = fill

	# fill Text dbfield

	page_count = 1
	pattern = r'^{{(?P<name>\w+):(?P<range>\d+)}}$'
	non_db_field = ['page_num', 'page_total']
	rich_db_field = ['service_provided', 'parts_replaced', 'remark', 'conclusion']
	thin_border = xlsx_Border(
		left=xlsx_Side(style='thin'), right=xlsx_Side(style='thin'), 
		top=xlsx_Side(style='thin'), bottom=xlsx_Side(style='thin'))
	bottom_border = xlsx_Border(
		left=None, right=None, 
		top=None, bottom=xlsx_Side(style='thin'))

	# if Record not complete, redirect to Edit page
	if ( q_pg.department_item == None ):
		redirect_url = reverse( 'photologue:photogroup-pm-edit', args=[q_pg.pk] )
		return redirect( redirect_url )

	# Fill values from database
	for cell in ws.get_cell_collection():
		if cell.value:
			res = re.match( pattern, unicode(cell.value) )
			if res:
				db_field = res.group('name')
				if db_field not in non_db_field:
					q_db_field = eval( "q_pg.{}".format( db_field ) )
					if type(q_db_field) == timezone.datetime:
						db_value = q_db_field.strftime( "%Y-%m-%d" )
					elif db_field == "company":
						db_value = q_pg.department_item.department.company.name
					elif db_field == "department":
						db_value = q_pg.department_item.department.name
					elif db_field == "department_item":
						db_value = q_pg.department_item.name
					elif db_field in rich_db_field:
						cell.alignment = xlsx_alignment(wrapText=True, vertical='top') 
						db_value = q_db_field
					else:
						db_value = q_db_field
				else:
					if db_field == "page_num":
						db_value = "Page : {}".format( page_count );
						page_count += 1
					else:
						print( 'non db_value FIXME {}'.format( db_field ) )
						db_value = "FIXME"
				cell.value = db_value
				#cell.border = thin_border
				#for i in range(0,int(res.group('range'))):
				#	print( u"cell offset {}:{}".format(cell.offset( row=i).column , cell.offset( row=i).row) )
				#	cell.offset( row=i ).border = thin_border
				cell_offset = int(res.group('range'))-1
				if cell_offset > 0:
					cell_range = "{}{}:{}{}".format(
									 cell.column, cell.row,
									 cell.offset( column=cell_offset ).column,
									 cell.row )
					style_range( ws, cell_range, border=bottom_border )

	# PM template merged cell style fix for openpyxl
	if 'PM' in str(q_pg.record_type).upper():
		style_cells_bottom = { 
			'B5:C5', 'F5:G5', 'B6:C6', 'F6:G6', 'F7:G7',
			'F29:G29', 
			}
		style_cells_square = { 
			'B9:E9', 'F9:G9',
			'B10:E10', 'B11:E11', 'B12:E12', 'B13:E13', 'B14:E14', 'B15:E15', 'B16:E16', 'B17:E17',
			'B18:E18', 'A19:G19', 'B20:E20', 'B21:E21', 'B22:E22', 'B23:E23',
			'F10:G10', 'F11:G11', 'F12:G12', 'F13:G13', 'F14:G14',
			'F15:G15', 'F16:G16', 'F17:G17', 'F18:G18', 
			'F20:G20', 'F21:G21', 'F22:G22', 'F23:G23', 
			}
		for style_cell in style_cells_square:
			style_range( ws, style_cell, border=thin_border )
		for style_cell in style_cells_bottom:
			style_range( ws, style_cell, border=bottom_border )
	               

	# Insert LOGO for text report xlsx
	img_report = create_image( logo_url, height=ws_photo['logo']['height'])
	ws.add_image( img_report, "{}1".format(ws_photo['logo']['record']) )
	
	# get all images 
	qset_photorecord = q_pg.photo_records.all() 
	qset_photos_before = q_pg.photo_records.filter( photo_class__name = "Before" )
	qset_photos_center = q_pg.photo_records.filter( photo_class__name = "Center" )
	qset_photos_after  = q_pg.photo_records.filter( photo_class__name = "After" )

	# insert photos
	len_before = len(qset_photos_before)
	len_center = len(qset_photos_center)
	len_after  = len(qset_photos_after)
	print( "debug: len_qset: {}/{}/{}".format( len_before, len_center, len_after ) )
	photo_scale = 2.25

	# no center
	if (len_before > 0) & (len_center == 0) & (len_after > 0):
		ws_photo['scale'] = 0.22
		ws_photo['x_scale'] = ws_photo['scale']
		ws_photo['y_scale'] = ws_photo['scale']

		ws_photo['col']['Before'] = 'A'
		ws_photo['col']['After'] = 'F'
		ws_photo['xoffset']['Before']  = 10
		ws_photo['xoffset']['After']  = 10

	# have left, center and right
	if (len_before > 0) & (len_center > 0) & (len_after > 0):
		ws_photo['scale'] = 0.18
		ws_photo['x_scale'] = ws_photo['scale']
		ws_photo['y_scale'] = ws_photo['scale']

		ws_photo['col']['Before'] = 'A'
		ws_photo['col']['Center'] = 'D'
		ws_photo['col']['After'] = 'G'

	# -=-==--=-=-=-=-=-=-=-==-=-
	# XlsxWriter photo workbook
	ws_writer = wb_writer.add_worksheet()
	
	# set print margin and paper etc
	ws_writer.set_paper( ws_photo['paper']  ) # A4 paper http://xlsxwriter.readthedocs.io/page_setup.html 
	ws_writer.set_margins( 
		left = ws_photo['margin']['left'], right = ws_photo['margin']['right'], 
		top = ws_photo['margin']['top'], bottom = ws_photo['margin']['bottom'] )

	# FIXME XlsxWriter set_column does not work in google sheet, 
	#ws_writer.set_column( 0, 100, ws_photo['cell']['width'] )
	#cell_format = wb_writer.add_format({'bold':True})
	#ws_writer.set_column( 'A:Z', ws_photo['cell']['width'], cell_format )

	# reszie entire worksheet
	page = {
		'row':{'Before':ws_photo['row_begin'], 'Center':ws_photo['row_begin'], 'After':ws_photo['row_begin']}, 
		'count':{'Before':0, 'Center':0, 'After':0},
		}

	# compute height from aspect ratio of photo
	def get_height( photo, scale ):
		org_height = float(photo.image.height)
		org_width = float(photo.image.width)
		display_width = 1024.0
		display_height = org_height / org_width * 1024.0
		new_height = display_height * scale / float(ws_photo['cell']['height'])
		return new_height 

	# Juicy Image Gallery
	for q_photorecord in qset_photorecord:
		photo_class = q_photorecord.photo_class.name 
		photo = q_photorecord.photo
		#url = app_url + q_photorecord.photo.image.url
		url = app_url + q_photorecord.photo.get_display_url()
		img_data = create_image_data( url )  
		next_photo_row = page['row'][photo_class] + get_height( photo, ws_photo['scale'] ) + ws_photo['gap']
		print( "{}:next_photo_row={}".format( photo.pk, next_photo_row ) )
		if next_photo_row > (ws_photo['page_col'] * (page['count'][photo_class]+1)):
			page['count'][photo_class] += 1
			print ("[next page]:{}:{}:{}".format( page['count'][photo_class], photo_class, page['row'][photo_class] ) )
			page['row'][photo_class] = ws_photo['page_col'] * page['count'][photo_class] + \
			                           ws_photo['row_begin']
		cell_name = "{}{}".format( ws_photo['col'][photo_class], page['row'][photo_class] )
		ws_writer.insert_image(cell_name, url, {
			'image_data':img_data, 'x_scale': ws_photo['x_scale'], 'y_scale': ws_photo['y_scale'],
			'x_offset': ws_photo['xoffset'][photo_class], 'y_offset': 0,
			} ) 
		page['row'][photo_class] += get_height( photo, ws_photo['scale'] ) + ws_photo['gap']
		# print( '{}:{}'.format( photo_class, page['row'][photo_class] ) ) 

	page_max = max( list( (page['count']['Before'], 
	                      page['count']['Center'],
						  page['count']['After'] )) )
	
	# insert new page header and Logo 
	# PM
	if 'CM' in str(q_pg.record_type).upper():
		print( "will add logo to CM page_max={}".format(page_max) )
		img_data = create_image_data( logo_url )

		# json load from cm-config.json
		json_config = json.loads( unicode( open( fn_config, 'r').read(), 'utf8' ))
		headers = json_config['cm_headers']
		print( 'headers={}'.format(headers) )
		headers[4]['text'] = str(q_pg.serial_no)
		headers[5]['text'] = str(q_pg.date_of_service.date())

		for page in range(0,page_max+1):
			cell_name = '{}{}'.format( ws_photo['logo']['photo'],
			                           page * ws_photo['page_row'] + 1 )
			ws_writer.insert_image( cell_name, logo_url, {
				'image_data':img_data, 'x_scale': ws_photo['logo']['x_scale'], 
				'y_scale': ws_photo['logo']['y_scale'] } )
			# will insert text and page num
			for header in headers: 
				merge_row = page * ws_photo['page_row'] + header['cell']['row']
				merge_format = wb_writer.add_format({ 'align':header['align'] } )
				merge_cell = "{}{}:{}{}".format( header['cell']['left'], merge_row,
				                                 header['cell']['right'], merge_row)
				#ws_writer.write( header['cell'], header['text'] )	
				ws_writer.merge_range( merge_cell, header['text'], merge_format)

	if 'PM' in str(q_pg.record_type).upper():
		img_data = create_image_data( logo_url )
		for page in range(0,page_max+1):
			cell_name = '{}{}'.format( ws_photo['logo']['photo'],
			                           page * ws_photo['page_row'] + 1 )
			ws_writer.insert_image( cell_name, logo_url, {
				'image_data':img_data, 'x_scale': ws_photo['logo']['x_scale'], 
				'y_scale': ws_photo['logo']['y_scale'] } )
			# will insert text and page num
		
	# insert new page header and Logo
	# CM
	if 'CM' in str(q_pg.record_type).upper():
		print( "will add logo to CM page_max={}".format(page_max) )
					
	# save and exit

	print( 'saving {}'.format(fn_photo_out))
	wb_text.save( fn_text_out )
	wb_photo.save( fn_photo_out )
	wb_writer.close()

	# create zipfile
	fn_xlsx_zip = os.path.join( tmp_root, filename_zip_out )
	xlsx_zip = zipfile.ZipFile( fn_xlsx_zip,'w' )
	xlsx_zip.write( fn_text_out,  filename_text_out )
	xlsx_zip.write( fn_writer_out,  filename_writer_out ) 
	counter = 1
	for q_pr in qset_photorecord:
		fullpath = q_pr.photo.image.path
		path, ext = os.path.splitext( fullpath )
		filename = "{}_{:02d}.{}".format( filename_serial , counter , ext )
		counter += 1
		xlsx_zip.write( q_pr.photo.image.path, filename )
	xlsx_zip.close()

	# HTML response 
	fn_io = StringIO( open(fn_xlsx_zip).read() )
	wrapper = FileWrapper( fn_io, blksize=5 )
	response = StreamingHttpResponse( wrapper, 
	               content_type = mimetypes.guess_type(fn_xlsx_zip)[0])
	response['Content-Disposition'] = "attachment; filename={}".format( filename_zip_out )
	return response

def GenerateXLSXAll(request):
	
	qset_photogroup = PhotoGroup.objects.all()
	print( 'debug: Generating {} xlsx'.format(len(qset_photogroup)))
	for q_photogroup in qset_photogroup:
		#reverse( 'photologue:generate-xlsx', args=[ q_photogroup.pk ] )
		response = GenerateXLSX( request, q_photogroup.pk )
		try:
			print( 'response content = {}'.format(
				response['Content-Disposition'] ) )
		except:
			print( 'ERROR: {}'.format( q_photogroup.pk ) )

	return redirect( reverse( 'photologue:monthly-report-list' )  )

class InventoryListView( ListView ):
	template_name = 'photologue/inventory-list.html'
	model = InventoryType 

class InventoryTypeDetail( DetailView ):
	model = InventoryType
	form_class = InventoryTypeForm

class InventoryTypeUpdate( UpdateView ):
	model = InventoryType
	form_class = InventoryTypeForm

class InventoryItemDetail( DetailView ):
	model = InventoryItem
	form_class = InventoryItemForm

class InventoryItemUpdate( UpdateView ):
	model = InventoryItem
	form_class = InventoryItemForm

def InventoryCheckout( request ):
	qset = request.POST.getlist('delete_photo') 
	print( "inventory checkout {}".format( qset ))

	for q_pk in qset:
		#q_photo = Photo.objects.get( pk=q_pk )
		#q_photo.checkout = True
		#q_photo.date_checkout = timezone.localtime()
		#q_photo.save()
		q_item = InventoryItem.objects.get( pk = q_pk )
		q_item.checked_out = True
		q_item.checkout_datetime = timezone.localtime()
		q_item.save() 

	return redirect( reverse( 'photologue:inventory-list' ) )

def InventorySet( request, photo_pk ):
	q_photo = Photo.objects.get( pk = photo_pk )
	print( u"{}/{}".format( request.POST, photo_pk  ) )
	inventory_pk = request.POST['inventory']
	#q_photo.department_item = DepartmentItem.objects.get( name = "Inventory" )
	#q_photo.inventory_type = InventoryType.objects.get( pk=inventory_pk )
	#q_photo.checkout = False
	inventory_type = InventoryType.objects.get( pk = inventory_pk )
	dt_added = timezone.localtime()
	path, filename = os.path.split(q_photo.image.path)
	dt, ext = os.path.splitext( filename )
	#FIXME use regexp instead of split
	serial_no = dt.split('-')[6] 
	item_name = u"{}_{}".format( 
			inventory_type.name.replace(" ","_"),
			dt_added.strftime("%y%m%d_%H%M%S") )

	print( u"serial:{} name:{} type:{}".format(
		serial_no, item_name, inventory_type.name ) )

	q_inventoryitem = InventoryItem( 
		name = item_name, serial_no = serial_no,
		)
	q_inventoryitem.inventory_type = inventory_type
	q_inventoryitem.save()
	q_inventoryitem.photos.add( q_photo ) 

	return redirect( reverse( 'photologue:inventory-list') )

class InventoryTypeCreate( CreateView ):
	model = InventoryType
	fields = ['name', 'description']


# AJAX form create views
class PhotoGroupCMView( UpdateView ):
	template_name = 'photologue/photogroup-pm-edit.html'
	model = PhotoGroup
	form_class = PhotoGroupCMForm

class PhotoGroupPMView( UpdateView ):
	template_name = 'photologue/photogroup-pm-edit.html' 
	model = PhotoGroup
	form_class = PhotoGroupPMForm 

def SetActivePhotoGroupView( request, pk ):
	print("set active photo group")
	profile = request.user.profile
	profile.active_photogroup = PhotoGroup.objects.get(pk=pk)
	profile.save()
	return redirect( reverse( 'photologue:monthly-report-detail', args=[pk] ) )

def AddPhotoActivePhotoGroupView( request, photo_pk ):
	print("add photo to active photo group" ) 

	profile = request.user.profile
	active_photogroup = profile.active_photogroup
	pg_image_class = PhotoGroupImageClass.objects.get( name = "Center" )
	q_photo = Photo.objects.get( pk = photo_pk )
	pg_image = PhotoGroupImage( photo = q_photo, photo_class = pg_image_class )
	pg_image.save()
	active_photogroup.photo_records.add( pg_image )
	return redirect( reverse( 'photologue:monthly-report-detail', args=[active_photogroup.pk] ) )

# Gallery views.

class GalleryListView(ListView):
	queryset = Gallery.objects.on_site().is_public()
	paginate_by = 20


class GalleryDetailView(DetailView):
	queryset = Gallery.objects.on_site().is_public()


class GalleryDateView(object):
	queryset = Gallery.objects.on_site().is_public()
	date_field = 'date_added'
	allow_empty = True


class GalleryDateDetailView(GalleryDateView, DateDetailView):
	pass


class GalleryArchiveIndexView(GalleryDateView, ArchiveIndexView):
	pass


class GalleryDayArchiveView(GalleryDateView, DayArchiveView):
	pass


class GalleryMonthArchiveView(GalleryDateView, MonthArchiveView):
	pass


class GalleryYearArchiveView(GalleryDateView, YearArchiveView):
	make_object_list = True

# Photo views.


class PhotoListView(ListView):
	queryset = Photo.objects.on_site().is_public()
	paginate_by = 20


class PhotoDetailView(DetailView):
	queryset = Photo.objects.on_site().is_public()

	def get_context_data( self, **kwargs):
		context = super( PhotoDetailView, self).get_context_data(**kwargs)
		profile = self.request.user.profile
		context['inventory_type_list'] = InventoryType.objects.all()

		return context


class PhotoDateView(object):
	queryset = Photo.objects.on_site().is_public()
	date_field = 'date_added'
	allow_empty = True


class PhotoDateDetailView(PhotoDateView, DateDetailView):
	pass


class PhotoArchiveIndexView(PhotoDateView, ArchiveIndexView):
	pass


class PhotoDayArchiveView(PhotoDateView, DayArchiveView):
	pass


class PhotoMonthArchiveView(PhotoDateView, MonthArchiveView):
	pass


class PhotoYearArchiveView(PhotoDateView, YearArchiveView):
	make_object_list = True


# Deprecated views.

class DeprecatedMonthMixin(object):

	"""Representation of months in urls has changed from a alpha representation ('jan' for January)
	to a numeric representation ('01' for January).
	Properly deprecate the previous urls."""

	query_string = True

	month_names = {'jan': '01',
				   'feb': '02',
				   'mar': '03',
				   'apr': '04',
				   'may': '05',
				   'jun': '06',
				   'jul': '07',
				   'aug': '08',
				   'sep': '09',
				   'oct': '10',
				   'nov': '11',
				   'dec': '12', }

	def get_redirect_url(self, *args, **kwargs):
		print('a')
		warnings.warn(
			DeprecationWarning('Months are now represented in urls by numbers rather than by '
							   'their first 3 letters. The old style will be removed in Photologue 3.4.'))


class GalleryDateDetailOldView(DeprecatedMonthMixin, RedirectView):
	permanent = True

	def get_redirect_url(self, *args, **kwargs):
		super(GalleryDateDetailOldView, self).get_redirect_url(*args, **kwargs)
		return reverse('photologue:gallery-detail', kwargs={'year': kwargs['year'],
															'month': self.month_names[kwargs['month']],
															'day': kwargs['day'],
															'slug': kwargs['slug']})


class GalleryDayArchiveOldView(DeprecatedMonthMixin, RedirectView):
	permanent = True

	def get_redirect_url(self, *args, **kwargs):
		super(GalleryDayArchiveOldView, self).get_redirect_url(*args, **kwargs)
		return reverse('photologue:gallery-archive-day', kwargs={'year': kwargs['year'],
																 'month': self.month_names[kwargs['month']],
																 'day': kwargs['day']})


class GalleryMonthArchiveOldView(DeprecatedMonthMixin, RedirectView):
	permanent = True

	def get_redirect_url(self, *args, **kwargs):
		super(GalleryMonthArchiveOldView, self).get_redirect_url(*args, **kwargs)
		return reverse('photologue:gallery-archive-month', kwargs={'year': kwargs['year'],
																   'month': self.month_names[kwargs['month']]})


class PhotoDateDetailOldView(DeprecatedMonthMixin, RedirectView):
	permanent = True

	def get_redirect_url(self, *args, **kwargs):
		super(PhotoDateDetailOldView, self).get_redirect_url(*args, **kwargs)
		return reverse('photologue:photo-detail', kwargs={'year': kwargs['year'],
														  'month': self.month_names[kwargs['month']],
														  'day': kwargs['day'],
														  'slug': kwargs['slug']})


class PhotoDayArchiveOldView(DeprecatedMonthMixin, RedirectView):
	permanent = True

	def get_redirect_url(self, *args, **kwargs):
		super(PhotoDayArchiveOldView, self).get_redirect_url(*args, **kwargs)
		return reverse('photologue:photo-archive-day', kwargs={'year': kwargs['year'],
															   'month': self.month_names[kwargs['month']],
															   'day': kwargs['day']})


class PhotoMonthArchiveOldView(DeprecatedMonthMixin, RedirectView):
	permanent = True

	def get_redirect_url(self, *args, **kwargs):
		super(PhotoMonthArchiveOldView, self).get_redirect_url(*args, **kwargs)
		return reverse('photologue:photo-archive-month', kwargs={'year': kwargs['year'],
																 'month': self.month_names[kwargs['month']]})
